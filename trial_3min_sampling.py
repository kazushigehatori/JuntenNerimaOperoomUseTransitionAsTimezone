"""3分区間サンプリングによる稼働率計算（試行）"""
import openpyxl
import datetime as dt

INPUT_FILE = "時間帯別稼働推移元データ.xlsx"
wb = openpyxl.load_workbook(INPUT_FILE)


def to_minutes(t):
    if isinstance(t, dt.time):
        return t.hour * 60 + t.minute
    elif isinstance(t, dt.timedelta):
        return int(t.total_seconds()) // 60
    elif isinstance(t, str):
        parts = t.split(":")
        return int(parts[0]) * 60 + int(parts[1])
    else:
        return t.hour * 60 + t.minute


# 定義シート読み込み
ws_def = wb["定義"]
room_weight = {}
for row in ws_def.iter_rows(min_row=2, max_row=20, min_col=1, max_col=2, values_only=True):
    if row[0] is not None and row[1] is not None:
        try:
            room_weight[str(row[0])] = float(row[1])
        except:
            break

exclude_weekdays = set()
r = 14
while True:
    r += 1
    v = ws_def.cell(row=r, column=1).value
    if v is None or v == "":
        break
    exclude_weekdays.add(str(v))

# データ読み込み
ws_data = wb["時間帯別稼働推移元データ"]
records = []
for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, values_only=True):
    mgmt_no, op_date, weekday, room, start_time, end_time, category = row
    if room is not None:
        records.append({
            "date": str(op_date) if op_date else "",
            "weekday": str(weekday) if weekday else "",
            "room": str(room),
            "start": start_time,
            "end": end_time,
            "category": str(category) if category else "",
        })

# フィルタ: 除外曜日除去（土日）、対象室のみ、全手術
filtered = [r for r in records if r["weekday"] not in exclude_weekdays
            and r["room"] in room_weight]

# 3分区間の生成: 9:00-9:02, 9:03-9:05, ..., 16:57-16:59
# 9:00 = 540分, 16:57 = 1017分 → 区間開始: 540, 543, 546, ..., 1017
# 区間数 = (1017 - 540) / 3 + 1 = 477/3 + 1 = 159 + 1 = 160
INTERVAL_MINUTES = 3
START_MIN = 9 * 60      # 540
END_MIN = 16 * 60 + 59  # 1019

intervals = []  # (start_min, end_min) のリスト
m = START_MIN
while m + INTERVAL_MINUTES - 1 <= END_MIN:
    intervals.append((m, m + INTERVAL_MINUTES - 1))
    m += INTERVAL_MINUTES

num_intervals = len(intervals)

# 日別にグループ化
days = {}
for r in filtered:
    d = r["date"]
    if d not in days:
        days[d] = []
    days[d].append(r)

num_days = len(days)

# ウェイト合計
WEIGHT_SUM = sum(room_weight.values())

# 計算: 各日・各区間・各部屋で「使用中」かを判定
total_weighted_usage = 0.0

for day_str, day_records in days.items():
    for iv_start, iv_end in intervals:
        # 各部屋について、この区間に重なる手術があるか判定
        for rm, weight in room_weight.items():
            in_use = False
            for r in day_records:
                if r["room"] != rm:
                    continue
                op_start = to_minutes(r["start"])
                op_end = to_minutes(r["end"])
                # 区間 [iv_start, iv_end] と手術 [op_start, op_end] が重なるか
                if op_start <= iv_end and iv_start <= op_end:
                    in_use = True
                    break  # 同一部屋・同一区間で1回だけカウント
            if in_use:
                total_weighted_usage += weight

# 分母
denominator = WEIGHT_SUM * num_intervals * num_days

# 稼働率
utilization_rate = total_weighted_usage / denominator * 100

# 出力
print(f"=== 3分区間サンプリング計算 ===")
print(f"対象期間: 2025年9月")
print(f"平日日数: {num_days}日")
print(f"時間帯: 9:00-16:59")
print(f"区間数: {num_intervals}（3分刻み）")
print(f"対象手術: 全手術 {len(filtered)}件")
print(f"部屋数: {len(room_weight)}室（ウェイト合計{WEIGHT_SUM}）")
print(f"使用ウェイト合計: {total_weighted_usage:.1f}")
print(f"分母: {denominator:.1f}")
print(f"稼働率: {utilization_rate:.1f}%")

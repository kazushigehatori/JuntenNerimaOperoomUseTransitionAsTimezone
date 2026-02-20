"""
HOGY社方式との差異調査 — 全30試行の稼働率比較
==============================================
弊社（メドライン）計算: 76.0%
HOGY社計算: 67.9%
差: 8.1ポイント

各試行で異なる条件の稼働率を算出し、67.9%に最も近い組み合わせを探索する。
"""

import openpyxl
import datetime as dt
import os
import sys

if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, "時間帯別稼働推移元データ.xlsx")


def to_minutes(t):
    if isinstance(t, dt.time):
        return t.hour * 60 + t.minute
    elif isinstance(t, dt.timedelta):
        return int(t.total_seconds()) // 60
    elif isinstance(t, str):
        parts = t.split(":")
        return int(parts[0]) * 60 + int(parts[1])
    else:
        return t.hour * 60 + t.minute


def make_slots(start_h, start_m, end_h, end_m):
    """指定範囲の30分刻みスロットを生成"""
    slots = []
    m = start_h * 60 + start_m
    end = end_h * 60 + end_m
    while m <= end:
        slots.append(dt.time(m // 60, m % 60))
        m += 30
    return slots


def main():
    print(f"入力ファイル: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE)

    # --- 定義シートから設定読み込み ---
    ws_def = wb["定義"]
    room_weight = {}
    for row in ws_def.iter_rows(min_row=2, max_row=20, min_col=1, max_col=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            try:
                room_weight[str(row[0])] = float(row[1])
            except (ValueError, TypeError):
                break
    print(f"対象手術室: {room_weight}")
    weight_total = sum(room_weight.values())
    print(f"ウェイト合計: {weight_total}")

    exclude_weekdays = set()
    r = 14
    while True:
        r += 1
        v = ws_def.cell(row=r, column=1).value
        if v is None or v == "":
            break
        exclude_weekdays.add(str(v))

    # --- 元データ読み込み ---
    ws_data = wb["時間帯別稼働推移元データ"]
    records = []
    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, values_only=True):
        mgmt_no, op_date, weekday, room, start_time, end_time, category = row
        if room is not None:
            records.append({
                "date": str(op_date) if op_date else "",
                "weekday": str(weekday) if weekday else "",
                "room": str(room),
                "start": start_time,
                "end": end_time,
                "category": str(category) if category else "",
            })
    wb.close()

    records_filtered = [r for r in records if r["weekday"] not in exclude_weekdays]
    all_dates = sorted(set(r["date"] for r in records_filtered))
    num_days = len(all_dates)
    print(f"対象レコード数: {len(records_filtered)}, 対象日数: {num_days}")

    cat_counts = {}
    for r in records_filtered:
        cat_counts[r["category"]] = cat_counts.get(r["category"], 0) + 1
    print(f"区分別件数: {cat_counts}")

    num_rooms = len(room_weight)

    # --- スロット定義 ---
    slots_16 = make_slots(9, 0, 16, 30)       # 9:00〜16:30 = 16区間
    slots_14 = make_slots(9, 0, 15, 30)       # 9:00〜15:30 = 14区間 (〜16:00)
    slots_18_early = make_slots(8, 30, 17, 0)  # 8:30〜17:00 = 18区間
    slots_18_late = make_slots(9, 0, 17, 30)   # 9:00〜17:30 = 18区間
    print(f"スロット: 16区間={len(slots_16)}, 14区間={len(slots_14)}, "
          f"18区間(8:30-)={len(slots_18_early)}, 18区間(9:00-)={len(slots_18_late)}")

    # --- データセット ---
    all_surgery = [r for r in records_filtered if r["room"] in room_weight]
    scheduled_only = [r for r in all_surgery if r["category"] == "定時"]
    room_weight_flat = {room: 1.0 for room in room_weight}
    weight_total_flat = sum(room_weight_flat.values())
    room_weight_no_angio = {k: v for k, v in room_weight.items() if k != "ｱﾝｷﾞｵ"}
    weight_no_angio = sum(room_weight_no_angio.values())
    scheduled_no_angio = [r for r in scheduled_only if r["room"] in room_weight_no_angio]

    # ========== 汎用計算関数 ==========
    def calc_overlap(data, slots, weights, weight_sum, offset_a, offset_b):
        """区間重なり方式: [snap+offset_a, snap+offset_b] と手術時間の重なり判定"""
        numerator = 0.0
        denominator = weight_sum * num_days * len(slots)
        for d in all_dates:
            day_recs = [r for r in data if r["date"] == d]
            for snap in slots:
                snap_min = to_minutes(snap)
                ia = snap_min + offset_a
                ib = snap_min + offset_b
                count = 0.0
                for r in day_recs:
                    room = r["room"]
                    if room not in weights:
                        continue
                    s = to_minutes(r["start"])
                    e = to_minutes(r["end"])
                    if ia <= e and s <= ib:
                        count += weights[room]
                numerator += count
        return (numerator / denominator * 100) if denominator > 0 else 0.0

    def calc_overlap_fixed_denom(data, slots, weights, denom_rooms, offset_a, offset_b):
        """区間重なり方式 + 分母=固定部屋数"""
        numerator = 0.0
        denominator = denom_rooms * num_days * len(slots)
        for d in all_dates:
            day_recs = [r for r in data if r["date"] == d]
            for snap in slots:
                snap_min = to_minutes(snap)
                ia = snap_min + offset_a
                ib = snap_min + offset_b
                count = 0.0
                for r in day_recs:
                    room = r["room"]
                    if room not in weights:
                        continue
                    s = to_minutes(r["start"])
                    e = to_minutes(r["end"])
                    if ia <= e and s <= ib:
                        count += weights[room]
                numerator += count
        return (numerator / denominator * 100) if denominator > 0 else 0.0

    def calc_snapshot(data, slots, weights, weight_sum):
        """スナップショット方式: start ≤ snap < end"""
        numerator = 0.0
        denominator = weight_sum * num_days * len(slots)
        for d in all_dates:
            day_recs = [r for r in data if r["date"] == d]
            for snap in slots:
                snap_min = to_minutes(snap)
                count = 0.0
                for r in day_recs:
                    room = r["room"]
                    if room not in weights:
                        continue
                    s = to_minutes(r["start"])
                    e = to_minutes(r["end"])
                    if s <= snap_min < e:
                        count += weights[room]
                numerator += count
        return (numerator / denominator * 100) if denominator > 0 else 0.0

    def calc_snapshot_fixed_denom(data, slots, weights, denom_rooms):
        """スナップショット方式 + 分母=固定部屋数"""
        numerator = 0.0
        denominator = denom_rooms * num_days * len(slots)
        for d in all_dates:
            day_recs = [r for r in data if r["date"] == d]
            for snap in slots:
                snap_min = to_minutes(snap)
                count = 0.0
                for r in day_recs:
                    room = r["room"]
                    if room not in weights:
                        continue
                    s = to_minutes(r["start"])
                    e = to_minutes(r["end"])
                    if s <= snap_min < e:
                        count += weights[room]
                numerator += count
        return (numerator / denominator * 100) if denominator > 0 else 0.0

    def calc_overlap_merged_1ab(data, slots, denom_rooms, offset_a, offset_b,
                                room_1a="01A", room_1b="01B"):
        """区間重なり方式 + 1A/1B統合カウント
        1A or 1Bが使用中 → 1室、両方使用中 → やはり1室（上限1）
        他の部屋は各1室。分母=denom_rooms。
        """
        target_rooms = set(room_weight.keys())
        other_rooms = target_rooms - {room_1a, room_1b}
        numerator = 0.0
        denominator = denom_rooms * num_days * len(slots)
        for d in all_dates:
            day_recs = [r for r in data if r["date"] == d]
            for snap in slots:
                snap_min = to_minutes(snap)
                ia = snap_min + offset_a
                ib = snap_min + offset_b
                count = 0.0
                # 他の部屋: 各1室
                for r in day_recs:
                    room = r["room"]
                    if room not in other_rooms:
                        continue
                    s = to_minutes(r["start"])
                    e = to_minutes(r["end"])
                    if ia <= e and s <= ib:
                        count += 1.0
                # 1A/1B統合: どちらか使用=1, 両方=1
                ab_used = False
                for r in day_recs:
                    room = r["room"]
                    if room not in (room_1a, room_1b):
                        continue
                    s = to_minutes(r["start"])
                    e = to_minutes(r["end"])
                    if ia <= e and s <= ib:
                        ab_used = True
                        break
                if ab_used:
                    count += 1.0
                numerator += count
        return (numerator / denominator * 100) if denominator > 0 else 0.0

    # ========== 全30試行 ==========
    target = 67.9

    # trials: (名前, 区間方式, ウェイト, 対象手術, 時間帯, 稼働率)
    trials = []

    # --- 試行1〜14 (既存) ---
    r1 = calc_overlap(all_surgery, slots_16, room_weight, weight_total, -14, +15)
    trials.append(("試行1", "弊社(-14/+15)", "定義準拠", "全手術", "9:00-16:30", r1))

    r2 = calc_snapshot(all_surgery, slots_16, room_weight, weight_total)
    trials.append(("試行2", "SS(点判定)", "定義準拠", "全手術", "9:00-16:30", r2))

    r3 = calc_snapshot(scheduled_only, slots_16, room_weight, weight_total)
    trials.append(("試行3", "SS(点判定)", "定義準拠", "定時のみ", "9:00-16:30", r3))

    r4 = calc_snapshot(scheduled_only, slots_16, room_weight_flat, weight_total_flat)
    trials.append(("試行4", "SS(点判定)", "W1.0固定", "定時のみ", "9:00-16:30", r4))

    r5 = calc_snapshot(scheduled_no_angio, slots_16, room_weight_no_angio, weight_no_angio)
    trials.append(("試行5", "SS(点判定)", "定義準拠", "定時ｱﾝｷﾞｵ除", "9:00-16:30", r5))

    r6 = calc_overlap(scheduled_only, slots_16, room_weight, weight_total, -14, +15)
    trials.append(("試行6", "弊社(-14/+15)", "定義準拠", "定時のみ", "9:00-16:30", r6))

    r7 = r2  # = 試行2
    trials.append(("試行7", "SS(点判定)", "定義準拠", "全手術", "9:00-16:30", r7))

    # 試行8: SS + 分母=実部屋数
    r8 = calc_snapshot_fixed_denom(all_surgery, slots_16, room_weight, num_rooms)
    trials.append(("試行8", "SS(点判定)", "分母=実部屋数", "全手術", "9:00-16:30", r8))

    r9 = calc_overlap(scheduled_only, slots_16, room_weight, weight_total, 0, +29)
    trials.append(("試行9", "HOGY(0/+29)", "定義準拠", "定時のみ", "9:00-16:30", r9))

    r10 = calc_overlap(all_surgery, slots_16, room_weight, weight_total, 0, +29)
    trials.append(("試行10", "HOGY(0/+29)", "定義準拠", "全手術", "9:00-16:30", r10))

    r11 = r9  # = 試行9
    trials.append(("試行11", "HOGY(0/+29)", "定義準拠", "定時のみ", "9:00-16:30", r11))

    r12 = r6  # = 試行6
    trials.append(("試行12", "弊社(-14/+15)", "定義準拠", "定時のみ", "9:00-16:30", r12))

    r13 = r10  # = 試行10
    trials.append(("試行13", "HOGY(0/+29)", "定義準拠", "全手術", "9:00-16:30", r13))

    r14 = r1  # = 試行1
    trials.append(("試行14", "弊社(-14/+15)", "定義準拠", "全手術", "9:00-16:30", r14))

    # --- 試行15〜26 (新規: HOGY区間固定・全手術で67.9%近似探索) ---

    # 試行15: HOGY区間 + 全手術 + 定義シートウェイト（= 試行10、ベースライン再確認）
    r15 = r10
    trials.append(("試行15", "HOGY(0/+29)", "定義準拠", "全手術", "9:00-16:30", r15))

    # 試行16: HOGY区間 + 全手術 + ウェイト1.0固定
    r16 = calc_overlap(all_surgery, slots_16, room_weight_flat, weight_total_flat, 0, +29)
    trials.append(("試行16", "HOGY(0/+29)", "W1.0固定", "全手術", "9:00-16:30", r16))

    # 試行17: HOGY区間 + 全手術 + 分母=固定部屋数（実部屋数、ウェイト無視）
    r17 = calc_overlap_fixed_denom(all_surgery, slots_16, room_weight, num_rooms, 0, +29)
    trials.append(("試行17", "HOGY(0/+29)", "分母=実部屋数", "全手術", "9:00-16:30", r17))

    # 試行18: HOGY区間 + 全手術 + 9:00-16:00（14区間）
    r18 = calc_overlap(all_surgery, slots_14, room_weight, weight_total, 0, +29)
    trials.append(("試行18", "HOGY(0/+29)", "定義準拠", "全手術", "9:00-16:00", r18))

    # 試行19: HOGY区間 + 全手術 + 8:30-17:00（18区間）
    r19 = calc_overlap(all_surgery, slots_18_early, room_weight, weight_total, 0, +29)
    trials.append(("試行19", "HOGY(0/+29)", "定義準拠", "全手術", "8:30-17:00", r19))

    # 試行20: HOGY区間 + 全手術 + 9:00-17:30（18区間）
    r20 = calc_overlap(all_surgery, slots_18_late, room_weight, weight_total, 0, +29)
    trials.append(("試行20", "HOGY(0/+29)", "定義準拠", "全手術", "9:00-17:30", r20))

    # 試行21: スナップショット + 全手術 + ウェイト1.0固定
    r21 = calc_snapshot(all_surgery, slots_16, room_weight_flat, weight_total_flat)
    trials.append(("試行21", "SS(点判定)", "W1.0固定", "全手術", "9:00-16:30", r21))

    # 試行22: スナップショット + 全手術 + 分母=固定部屋数
    r22 = calc_snapshot_fixed_denom(all_surgery, slots_16, room_weight, num_rooms)
    trials.append(("試行22", "SS(点判定)", "分母=実部屋数", "全手術", "9:00-16:30", r22))

    # 試行23: スナップショット + 全手術 + 9:00-16:00（14区間）
    r23 = calc_snapshot(all_surgery, slots_14, room_weight, weight_total)
    trials.append(("試行23", "SS(点判定)", "定義準拠", "全手術", "9:00-16:00", r23))

    # 試行24: HOGY区間 + 全手術 + ウェイト1.0 + 分母=固定部屋数
    r24 = calc_overlap_fixed_denom(all_surgery, slots_16, room_weight_flat, num_rooms, 0, +29)
    trials.append(("試行24", "HOGY(0/+29)", "W1.0+分母固定", "全手術", "9:00-16:30", r24))

    # 試行25: HOGY区間 + 全手術 + 部屋数を10室に変更（分母調整）
    r25 = calc_overlap_fixed_denom(all_surgery, slots_16, room_weight, 10, 0, +29)
    trials.append(("試行25", "HOGY(0/+29)", "分母=10室", "全手術", "9:00-16:30", r25))

    # 試行26: HOGY区間 + 全手術 + 部屋数を12室に変更（分母調整）
    r26 = calc_overlap_fixed_denom(all_surgery, slots_16, room_weight, 12, 0, +29)
    trials.append(("試行26", "HOGY(0/+29)", "分母=12室", "全手術", "9:00-16:30", r26))

    # --- 試行27〜30 (1A/1B統合方式) ---

    # 試行27: HOGY区間 + 全手術 + 1A/1B統合(使用=1,分母=10) + 9:00-16:30
    r27 = calc_overlap_merged_1ab(all_surgery, slots_16, 10, 0, +29)
    trials.append(("試行27", "HOGY(0/+29)", "1AB統合,分母10", "全手術", "9:00-16:30", r27))

    # 試行28: HOGY区間 + 全手術 + 1A/1B統合(使用=1,分母=9) + 9:00-16:30
    # ※分母9 = 1A/1Bで1室 + 残り8室
    r28 = calc_overlap_merged_1ab(all_surgery, slots_16, 9, 0, +29)
    trials.append(("試行28", "HOGY(0/+29)", "1AB統合,分母9", "全手術", "9:00-16:30", r28))

    # 試行29: HOGY区間 + 定時のみ + 1A/1B統合(使用=1,分母=10) + 9:00-16:30
    r29 = calc_overlap_merged_1ab(scheduled_only, slots_16, 10, 0, +29)
    trials.append(("試行29", "HOGY(0/+29)", "1AB統合,分母10", "定時のみ", "9:00-16:30", r29))

    # 試行30: HOGY区間 + 定時のみ + 1A/1B統合(使用=1,分母=9) + 9:00-16:30
    r30 = calc_overlap_merged_1ab(scheduled_only, slots_16, 9, 0, +29)
    trials.append(("試行30", "HOGY(0/+29)", "1AB統合,分母9", "定時のみ", "9:00-16:30", r30))

    # ========== 全試行を67.9%に近い順でソート ==========
    print(f"\n{'='*110}")
    print(f"=== 全試行結果（67.9%に近い順） ===")
    print(f"{'='*110}")
    print(f"{'順位':4s} {'試行':8s} {'区間方式':16s} {'ウェイト':14s} {'対象手術':12s} "
          f"{'時間帯':12s} {'稼働率':8s} {'差':8s}")
    print(f"{'-'*110}")

    sorted_trials = sorted(trials, key=lambda x: abs(x[5] - target))
    for rank, (name, method, weight, surgery, slots_desc, rate) in enumerate(sorted_trials, 1):
        diff = rate - target
        mark = " ★" if abs(diff) <= 2.0 else ""
        print(f"{rank:3d}  {name:8s} {method:16s} {weight:14s} {surgery:12s} "
              f"{slots_desc:12s} {rate:6.1f}% {diff:+6.1f}pt{mark}")

    print(f"\n目標: {target}%")
    best = sorted_trials[0]
    print(f"最も近い試行: {best[0]}（{best[5]:.1f}%、差 {abs(best[5]-target):.1f}pt）")
    print(f"条件: {best[1]} + {best[2]} + {best[3]} + {best[4]}")

    # ========== 新規試行15-26のみ抽出 ==========
    print(f"\n{'='*110}")
    print(f"=== 新規試行15-30（67.9%に近い順） ===")
    print(f"{'='*110}")
    print(f"{'順位':4s} {'試行':8s} {'区間方式':16s} {'ウェイト':14s} {'対象手術':12s} "
          f"{'時間帯':12s} {'稼働率':8s} {'差':8s}")
    print(f"{'-'*110}")

    new_trials = [(n, m, w, s, sl, r) for n, m, w, s, sl, r in trials
                  if int(n.replace("試行", "")) >= 15]
    new_sorted = sorted(new_trials, key=lambda x: abs(x[5] - target))
    for rank, (name, method, weight, surgery, slots_desc, rate) in enumerate(new_sorted, 1):
        diff = rate - target
        mark = " ★" if abs(diff) <= 2.0 else ""
        print(f"{rank:3d}  {name:8s} {method:16s} {weight:14s} {surgery:12s} "
              f"{slots_desc:12s} {rate:6.1f}% {diff:+6.1f}pt{mark}")

    # ========== 結論 ==========
    print(f"\n{'='*80}")
    print("=== 結論 ===")
    print(f"{'='*80}")
    print(f"全30試行中、最も67.9%に近い: {best[0]} = {best[5]:.1f}%（差{abs(best[5]-target):.1f}pt）")
    print(f"条件: {best[1]} + {best[2]} + {best[3]} + {best[4]}")

    # 上位3つ
    print("\n【上位3試行】")
    for i, (name, method, weight, surgery, slots_desc, rate) in enumerate(sorted_trials[:3], 1):
        diff = rate - target
        print(f"  {i}. {name}: {rate:.1f}% (差{diff:+.1f}pt) - {method} + {weight} + {surgery} + {slots_desc}")


if __name__ == "__main__":
    main()

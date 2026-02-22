"""
時間帯別稼働推移 計算スクリプト v4.0
====================================
手術実施データから、1分毎のサンプリング（8:00〜20:29）で各手術室の使用有無を判定し、
30分スナップショット（8:00〜20:00、25個）ごとに30個の1分サンプルの平均使用室数を算出します。

変更点（v4.0）:
- 計測区間方式（±14/+15分）→ 1分サンプリング + 30分平均方式
- 01Bの使用は01Aとしてカウント（01Bウェイト=0の場合、01Aに統合）
- 全室ウェイト1.0（定義シートから読み込み）

使い方:
    python calculate_timezone_usage.py

入力: 時間帯別稼働推移元データ.xlsx（同一フォルダに配置）
出力: 時間帯別稼働推移-結果.xlsx（同一フォルダに生成）
"""

import openpyxl
import datetime as dt
import os
import sys

# ========== 設定 ==========
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, "時間帯別稼働推移元データ.xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "時間帯別稼働推移-結果.xlsx")


def to_minutes(t):
    """時刻を分に変換（time, timedelta, str対応）"""
    if isinstance(t, dt.time):
        return t.hour * 60 + t.minute
    elif isinstance(t, dt.timedelta):
        return int(t.total_seconds()) // 60
    elif isinstance(t, str):
        parts = t.split(":")
        return int(parts[0]) * 60 + int(parts[1])
    else:
        return t.hour * 60 + t.minute


def main():
    print(f"入力ファイル読み込み: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE)

    # --- 定義シートから設定読み込み ---
    ws_def = wb["定義"]

    # 対象手術室とウェイト（全行読み込み）
    room_weight_raw = {}
    for row in ws_def.iter_rows(min_row=2, max_row=20, min_col=1, max_col=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            try:
                room_weight_raw[str(row[0])] = float(row[1])
            except (ValueError, TypeError):
                break

    # 01B統合ロジック: ウェイト0の部屋を特定し、統合先を決定
    # 01Bウェイト=0 → 01Bの手術データを01Aに統合
    merge_map = {}  # {統合元部屋名: 統合先部屋名}
    room_weight = {}  # 実際に集計に使うウェイト（ウェイト>0の部屋のみ）

    for room_name, weight in room_weight_raw.items():
        if weight == 0:
            # ウェイト0の部屋 → 統合先を探す
            # 01B → 01A のように、末尾アルファベット違いの部屋を探す
            base = room_name[:-1] if room_name[-1].isalpha() else None
            if base:
                for candidate, cw in room_weight_raw.items():
                    if candidate != room_name and candidate.startswith(base) and cw > 0:
                        merge_map[room_name] = candidate
                        break
            if room_name not in merge_map:
                print(f"警告: ウェイト0の部屋 '{room_name}' の統合先が見つかりません。無視します。")
        else:
            room_weight[room_name] = weight

    print(f"対象手術室: {room_weight}")
    if merge_map:
        print(f"部屋統合: {merge_map}")

    # 除外曜日
    exclude_weekdays = set()
    r = 14
    while True:
        r += 1
        v = ws_def.cell(row=r, column=1).value
        if v is None or v == "":
            break
        exclude_weekdays.add(str(v))
    print(f"除外曜日: {exclude_weekdays}")

    # --- 元データ読み込み ---
    ws_data = wb["時間帯別稼働推移元データ"]
    records = []
    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, values_only=True):
        mgmt_no, op_date, weekday, room, start_time, end_time, category = row
        if room is not None:
            room_str = str(room)
            # 部屋統合: 01B → 01A
            if room_str in merge_map:
                room_str = merge_map[room_str]
            records.append({
                "date": str(op_date) if op_date else "",
                "weekday": str(weekday) if weekday else "",
                "room": room_str,
                "start": start_time,
                "end": end_time,
                "category": str(category) if category else "",
            })

    print(f"総レコード数: {len(records)}")

    # 除外曜日フィルタリング
    records_filtered = [r for r in records if r["weekday"] not in exclude_weekdays]
    print(f"除外後レコード数: {len(records_filtered)}")

    # --- スナップショット時刻（8:00から30分おき、20:00まで = 25個）---
    snapshot_times = []
    for h in range(8, 20):
        snapshot_times.append(dt.time(h, 0))
        snapshot_times.append(dt.time(h, 30))
    snapshot_times.append(dt.time(20, 0))

    # --- 1分サンプリング + 30分平均集計関数 ---
    def count_rooms_at_snapshots(data):
        """
        各スナップショット時刻の30分間（+0〜+29分）について、
        1分毎のサンプリングで使用室数を計算し、30個の平均を日平均で算出。
        """
        days = {}
        for r in data:
            d = r["date"]
            if d not in days:
                days[d] = []
            days[d].append(r)

        num_days = len(days)
        if num_days == 0:
            return [0.0] * len(snapshot_times)

        totals = [0.0] * len(snapshot_times)

        for day_str, day_records in days.items():
            for si, snap in enumerate(snapshot_times):
                snap_min = to_minutes(snap)
                # 30分間の1分サンプリング: snap_min + 0, +1, ..., +29
                minute_sum = 0.0
                for offset in range(30):
                    sample_min = snap_min + offset
                    # 各部屋の使用有無を判定
                    room_used = set()
                    for r in day_records:
                        room = r["room"]
                        if room not in room_weight:
                            continue
                        start_min = to_minutes(r["start"])
                        end_min = to_minutes(r["end"])
                        # 使用中判定: 入室時刻 ≤ sample_min ≤ 麻酔終了時刻
                        if start_min <= sample_min <= end_min:
                            room_used.add(room)
                    # 使用室数 = 使用中の部屋のウェイト合計（各部屋上限1回）
                    count = sum(room_weight[rm] for rm in room_used)
                    minute_sum += count
                # 30分間の平均
                snapshot_avg = minute_sum / 30.0
                totals[si] += snapshot_avg

        averages = [round(t / num_days, 2) for t in totals]
        return averages

    # --- 全手術（定時・臨時・緊急）---
    all_surgery = [r for r in records_filtered if r["room"] in room_weight]
    print(f"\n全手術（対象室のみ）: {len(all_surgery)} 件")
    all_results = count_rooms_at_snapshots(all_surgery)

    # --- 予定手術のみ（定時のみ）---
    scheduled_only = [r for r in records_filtered if r["room"] in room_weight and r["category"] == "定時"]
    print(f"予定手術のみ: {len(scheduled_only)} 件")
    sched_results = count_rooms_at_snapshots(scheduled_only)

    # --- 計算結果シートに書き込み ---
    ws_result = wb["計算結果"]

    # 全体集計（Row2-3）
    for i, val in enumerate(all_results):
        ws_result.cell(row=2, column=2 + i, value=val)

    for i, val in enumerate(sched_results):
        ws_result.cell(row=3, column=2 + i, value=val)

    # --- 曜日別集計（Row5〜33）---
    weekday_rows = {
        "月曜日": {"all_row": 7,  "sched_row": 8},
        "火曜日": {"all_row": 12, "sched_row": 13},
        "水曜日": {"all_row": 17, "sched_row": 18},
        "木曜日": {"all_row": 22, "sched_row": 23},
        "金曜日": {"all_row": 27, "sched_row": 28},
        "土曜日": {"all_row": 32, "sched_row": 33},
    }

    print("\n--- 曜日別集計 ---")
    for weekday_name, rows in weekday_rows.items():
        weekday_records = [r for r in records if r["weekday"] == weekday_name and r["room"] in room_weight]
        weekday_scheduled = [r for r in weekday_records if r["category"] == "定時"]

        wd_all_results = count_rooms_at_snapshots(weekday_records)
        wd_sched_results = count_rooms_at_snapshots(weekday_scheduled)

        for i, val in enumerate(wd_all_results):
            ws_result.cell(row=rows["all_row"], column=2 + i, value=val)
        for i, val in enumerate(wd_sched_results):
            ws_result.cell(row=rows["sched_row"], column=2 + i, value=val)

        # 対象日数を算出
        wd_days = set(r["date"] for r in weekday_records)
        print(f"  {weekday_name}: 全手術={wd_all_results}, 予定={wd_sched_results}, 対象日数={len(wd_days)}")

    # --- 検証用シート（定時・臨時・緊急別の部屋数）---
    def count_rooms_by_day(data):
        """日別×スナップショット時刻の稼働室数（1分サンプリング30分平均）を返す: {date: [val, ...]}"""
        days = {}
        for r in data:
            d = r["date"]
            if d not in days:
                days[d] = []
            days[d].append(r)

        result = {}
        for day_str, day_records in sorted(days.items()):
            counts = []
            for si, snap in enumerate(snapshot_times):
                snap_min = to_minutes(snap)
                minute_sum = 0.0
                for offset in range(30):
                    sample_min = snap_min + offset
                    room_used = set()
                    for r in day_records:
                        room = r["room"]
                        if room not in room_weight:
                            continue
                        start_min = to_minutes(r["start"])
                        end_min = to_minutes(r["end"])
                        if start_min <= sample_min <= end_min:
                            room_used.add(room)
                    count = sum(room_weight[rm] for rm in room_used)
                    minute_sum += count
                snapshot_avg = round(minute_sum / 30.0, 4)
                counts.append(snapshot_avg)
            result[day_str] = counts
        return result

    # 区分別データ
    sched_data = [r for r in records_filtered if r["room"] in room_weight and r["category"] == "定時"]
    urgent_data = [r for r in records_filtered if r["room"] in room_weight and r["category"] == "臨時"]
    emerg_data = [r for r in records_filtered if r["room"] in room_weight and r["category"] == "緊急"]

    sched_by_day = count_rooms_by_day(sched_data)
    urgent_by_day = count_rooms_by_day(urgent_data)
    emerg_by_day = count_rooms_by_day(emerg_data)
    all_by_day = count_rooms_by_day(all_surgery)

    # 全日付の和集合（ソート済み）
    all_dates = sorted(set(list(sched_by_day.keys()) + list(urgent_by_day.keys()) +
                           list(emerg_by_day.keys()) + list(all_by_day.keys())))

    # 既存シートがあれば削除して再作成
    verify_sheet_name = "検証_定時臨時緊急別"
    if verify_sheet_name in wb.sheetnames:
        del wb[verify_sheet_name]
    ws_verify = wb.create_sheet(verify_sheet_name)

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font = Font(name="Meiryo UI", size=9, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2980B9")
    label_font = Font(name="Meiryo UI", size=10, bold=True)
    data_font = Font(name="Meiryo UI", size=9)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    num_weekdays = len(all_dates)  # 平日日数

    def write_section(ws, start_row, section_label, by_day_data, dates):
        """1セクションを書き込み、最終行+1を返す"""
        # セクションラベル
        cell = ws.cell(row=start_row, column=1, value=section_label)
        cell.font = label_font

        # ヘッダ行
        hr = start_row + 1
        dark_fill = PatternFill("solid", fgColor="1A5276")

        cell = ws.cell(row=hr, column=1, value="時間帯")
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

        # B列: 全平日合計ヘッダ
        cell = ws.cell(row=hr, column=2, value="全平日合計")
        cell.font = header_font
        cell.fill = dark_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

        # C列: 全平日平均ヘッダ
        cell = ws.cell(row=hr, column=3, value="全平日平均")
        cell.font = header_font
        cell.fill = dark_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

        # D列以降: 日付ヘッダ
        for di, d in enumerate(dates):
            cell = ws.cell(row=hr, column=4 + di, value=d)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # データ行
        bold_font = Font(name="Meiryo UI", size=9, bold=True)
        for si, snap in enumerate(snapshot_times):
            r = hr + 1 + si
            time_label = f"{snap.hour}:{snap.minute:02d}"
            cell = ws.cell(row=r, column=1, value=time_label)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            # B列: 全平日合計
            day_total = 0.0
            for d in dates:
                vals = by_day_data.get(d, [0.0] * len(snapshot_times))
                day_total += vals[si] if si < len(vals) else 0.0
            cell = ws.cell(row=r, column=2, value=day_total)
            cell.font = bold_font
            cell.border = thin_border
            cell.number_format = "0.0"

            # C列: 全平日平均
            avg = round(day_total / num_weekdays, 2) if num_weekdays > 0 else 0.0
            cell = ws.cell(row=r, column=3, value=avg)
            cell.font = bold_font
            cell.border = thin_border
            cell.number_format = "0.00"

            # D列以降: 日別値
            for di, d in enumerate(dates):
                vals = by_day_data.get(d, [0.0] * len(snapshot_times))
                val = vals[si] if si < len(vals) else 0.0
                cell = ws.cell(row=r, column=4 + di, value=val)
                cell.font = data_font
                cell.border = thin_border
                cell.number_format = "0.00"

        return hr + 1 + len(snapshot_times) + 1  # 次セクション開始行（1行空け）

    # 列幅設定
    ws_verify.column_dimensions["A"].width = 10
    ws_verify.column_dimensions["B"].width = 14
    ws_verify.column_dimensions["C"].width = 14
    for di in range(len(all_dates)):
        col_letter = openpyxl.utils.get_column_letter(4 + di)
        ws_verify.column_dimensions[col_letter].width = 12

    # 4セクション書き込み
    current_row = 1
    current_row = write_section(ws_verify, current_row, "【定時のみ】", sched_by_day, all_dates)
    current_row = write_section(ws_verify, current_row, "【臨時のみ】", urgent_by_day, all_dates)
    current_row = write_section(ws_verify, current_row, "【緊急のみ】", emerg_by_day, all_dates)
    current_row = write_section(ws_verify, current_row, "【合計（検証用）】", all_by_day, all_dates)

    print(f"\n検証シート '{verify_sheet_name}' を作成しました")

    # --- 検証：定時+臨時+緊急 ≒ 全手術か ---
    mismatch_count = 0
    total_cells = 0
    for di, d in enumerate(all_dates):
        s_vals = sched_by_day.get(d, [0.0] * len(snapshot_times))
        u_vals = urgent_by_day.get(d, [0.0] * len(snapshot_times))
        e_vals = emerg_by_day.get(d, [0.0] * len(snapshot_times))
        a_vals = all_by_day.get(d, [0.0] * len(snapshot_times))
        for si in range(len(snapshot_times)):
            total_cells += 1
            s = s_vals[si] if si < len(s_vals) else 0.0
            u = u_vals[si] if si < len(u_vals) else 0.0
            e = e_vals[si] if si < len(e_vals) else 0.0
            a = a_vals[si] if si < len(a_vals) else 0.0
            if abs((s + u + e) - a) > 0.01:
                mismatch_count += 1

    print(f"\n=== 検証結果 ===")
    print(f"1. 定時+臨時+緊急 vs 全手術 の一致チェック: "
          f"{'全セル一致 OK' if mismatch_count == 0 else f'{mismatch_count}/{total_cells} セル差異あり'}")
    if mismatch_count > 0:
        print(f"   (同一部屋・同一時間帯で異なる区分の手術が入れ替わる場合、"
              f"各区分別では各1回カウントされるが合計では1回のみのため差異が生じる。正常動作)")

    # 件数内訳
    print(f"\n2. 件数内訳:")
    print(f"   定時: {len(sched_data)} 件")
    print(f"   臨時: {len(urgent_data)} 件")
    print(f"   緊急: {len(emerg_data)} 件")
    print(f"   合計: {len(sched_data) + len(urgent_data) + len(emerg_data)} 件 "
          f"(全手術={len(all_surgery)} 件)")

    non_scheduled = len(urgent_data) + len(emerg_data)
    total = len(all_surgery)
    ratio = non_scheduled / total * 100 if total > 0 else 0
    print(f"\n3. 臨時+緊急が全体に占める割合: {non_scheduled}/{total} = {ratio:.1f}%")

    # --- サンプリング検証（20か所） ---
    import random
    random.seed(42)

    category_map = {
        "定時": sched_by_day,
        "臨時": urgent_by_day,
        "緊急": emerg_by_day,
        "合計": all_by_day,
    }
    category_filter = {
        "定時": lambda r: r["category"] == "定時",
        "臨時": lambda r: r["category"] == "臨時",
        "緊急": lambda r: r["category"] == "緊急",
        "合計": lambda r: True,
    }

    sample_dates = random.sample(all_dates, min(5, len(all_dates)))
    am_indices = [i for i, s in enumerate(snapshot_times) if s.hour < 12]
    pm_indices = [i for i, s in enumerate(snapshot_times) if 12 <= s.hour < 20]
    categories = ["定時", "臨時", "緊急", "合計"]

    samples = []
    for d in sample_dates:
        am_picks = random.sample(am_indices, min(2, len(am_indices)))
        pm_picks = random.sample(pm_indices, min(2, len(pm_indices)))
        for si in am_picks + pm_picks:
            cat = random.choice(categories)
            samples.append((d, si, cat))

    samples = samples[:20]

    print(f"\n=== サンプリング検証（{len(samples)}か所） ===")
    ok_count = 0
    for idx, (d, si, cat) in enumerate(samples):
        snap = snapshot_times[si]
        snap_min = to_minutes(snap)
        time_str = f"{snap.hour}:{snap.minute:02d}"

        # 検証シートの値
        by_day = category_map[cat]
        sheet_val = by_day.get(d, [0.0] * len(snapshot_times))[si]

        # 元データから手計算（1分サンプリング30分平均）
        cat_filter = category_filter[cat]
        day_records = [r for r in records_filtered
                       if r["date"] == d and r["room"] in room_weight and cat_filter(r)]
        minute_sum = 0.0
        for offset in range(30):
            sample_min = snap_min + offset
            room_used = set()
            for r in day_records:
                start_min = to_minutes(r["start"])
                end_min = to_minutes(r["end"])
                if start_min <= sample_min <= end_min:
                    room_used.add(r["room"])
            minute_sum += sum(room_weight[rm] for rm in room_used)
        calc_val = round(minute_sum / 30.0, 4)

        match = abs(sheet_val - calc_val) < 0.01
        status = "OK" if match else "NG"
        if match:
            ok_count += 1

        date_short = d.replace("2025/", "") if "2025/" in d else d
        print(f"#{idx+1:2d}  {date_short} {time_str} {cat:4s}  "
              f"検証シート={sheet_val:.4f}  手計算={calc_val:.4f}  {status}")

        if not match:
            print(f"     *** 不一致!")

    print(f"\n結果: {ok_count}/{len(samples)} 一致"
          f"（{ok_count/len(samples)*100:.0f}%）")

    # --- 最大値チェック ---
    weight_sum = sum(room_weight.values())
    max_val = 0.0
    max_info = ""
    over_count = 0
    for d in all_dates:
        for cat_name, by_day in [("定時", sched_by_day), ("臨時", urgent_by_day),
                                  ("緊急", emerg_by_day), ("合計", all_by_day)]:
            vals = by_day.get(d, [0.0] * len(snapshot_times))
            for si, v in enumerate(vals):
                if v > max_val:
                    max_val = v
                    snap = snapshot_times[si]
                    max_info = f"{d} {snap.hour}:{snap.minute:02d} {cat_name}"
                if v > weight_sum + 0.01:
                    over_count += 1
    print(f"\n=== 最大値チェック ===")
    print(f"ウェイト合計(上限): {weight_sum}")
    print(f"全セル最大値: {max_val:.4f} ({max_info})")
    print(f"上限超過セル数: {over_count}")

    # --- 別名保存 ---
    wb.save(OUTPUT_FILE)
    print(f"\n計算完了: {OUTPUT_FILE}")
    print(f"  全手術:   {all_results}")
    print(f"  予定のみ: {sched_results}")


if __name__ == "__main__":
    main()

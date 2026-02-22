"""他ソフトの分母28,980を再現する条件を推定する試行"""
import openpyxl
import datetime as dt

INPUT_FILE = "時間帯別稼働推移元データ.xlsx"
wb = openpyxl.load_workbook(INPUT_FILE)


def to_minutes(t):
    if isinstance(t, dt.time):
        return t.hour * 60 + t.minute
    elif isinstance(t, dt.timedelta):
        return int(t.total_seconds()) // 60
    elif isinstance(t, str):
        parts = t.split(":")
        return int(parts[0]) * 60 + int(parts[1])
    else:
        return t.hour * 60 + t.minute


# ========== データ読み込み ==========
ws_def = wb["定義"]
room_weight = {}
for row in ws_def.iter_rows(min_row=2, max_row=20, min_col=1, max_col=2, values_only=True):
    if row[0] is not None and row[1] is not None:
        try:
            room_weight[str(row[0])] = float(row[1])
        except:
            break

exclude_weekdays = set()
r = 14
while True:
    r += 1
    v = ws_def.cell(row=r, column=1).value
    if v is None or v == "":
        break
    exclude_weekdays.add(str(v))

ws_data = wb["時間帯別稼働推移元データ"]
records = []
for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, values_only=True):
    mgmt_no, op_date, weekday, room, start_time, end_time, category = row
    if room is not None:
        records.append({
            "date": str(op_date) if op_date else "",
            "weekday": str(weekday) if weekday else "",
            "room": str(room),
            "start": start_time,
            "end": end_time,
            "category": str(category) if category else "",
        })

filtered = [r for r in records if r["weekday"] not in exclude_weekdays
            and r["room"] in room_weight]

# 日別グループ化
days = {}
for r in filtered:
    d = r["date"]
    if d not in days:
        days[d] = []
    days[d].append(r)

num_days = len(days)

# ========== 1. 分母28,980の因数分解 ==========
print("=" * 60)
print("1. 分母 28,980 の因数分解")
print("=" * 60)
TARGET_DENOM = 28980
TARGET_USAGE = 19347
TARGET_RATE = 66.76

print(f"\n28,980 = {TARGET_DENOM}")
print(f"素因数分解: ", end="")
n = TARGET_DENOM
factors = []
for p in range(2, int(n**0.5) + 1):
    while n % p == 0:
        factors.append(p)
        n //= p
if n > 1:
    factors.append(n)
print(" x ".join(str(f) for f in factors))

print(f"\n20日固定で割り切れる組み合わせ (部屋数 x 区間数):")
print(f"{'部屋数':>8s}  {'区間数':>8s}  {'区間x3分':>10s}  {'時間':>8s}  備考")
print("-" * 60)
daily = TARGET_DENOM // num_days  # 1449
print(f"28,980 / 20日 = {daily}")
for rooms in range(1, 20):
    if daily % rooms == 0:
        intervals = daily // rooms
        total_min = intervals * 3
        hours = total_min // 60
        mins = total_min % 60
        note = ""
        if rooms == 9 and intervals == 161:
            note = "<-- 有力候補"
        elif rooms == 10:
            note = "(10室=1A,1B各1.0?)"
        print(f"{rooms:8d}  {intervals:8d}  {total_min:8d}分  {hours}h{mins:02d}m  {note}")

# ========== 2. 各仮説で計算 ==========
print("\n" + "=" * 60)
print("2. 各仮説での再計算")
print("=" * 60)

normal_rooms = ["02", "03", "05", "06", "07", "08", "09", "10"]


def calc_hypothesis(label, room_mode, num_rooms, interval_list):
    """
    room_mode:
      "weighted"   = 01A:0.5, 01B:0.5, 他:1.0 (弊社方式)
      "merged"     = 01A+01Bを1室統合(どちらか使用=1), 他:1.0 (9室)
      "separate"   = 01A:1.0, 01B:1.0, 他:1.0 (10室)
      "01A_only"   = 01Aのみ(1B除外), 他:1.0 (9室)
    """
    total = 0.0
    for day_str, day_records in days.items():
        for iv_start, iv_end in interval_list:
            # 通常部屋
            for rm in normal_rooms:
                in_use = False
                for r in day_records:
                    if r["room"] != rm:
                        continue
                    op_s = to_minutes(r["start"])
                    op_e = to_minutes(r["end"])
                    if op_s <= iv_end and iv_start <= op_e:
                        in_use = True
                        break
                if in_use:
                    total += 1.0

            # 01A/01B処理
            if room_mode == "weighted":
                for rm01 in ("01A", "01B"):
                    in_use = False
                    for r in day_records:
                        if r["room"] != rm01:
                            continue
                        op_s = to_minutes(r["start"])
                        op_e = to_minutes(r["end"])
                        if op_s <= iv_end and iv_start <= op_e:
                            in_use = True
                            break
                    if in_use:
                        total += 0.5

            elif room_mode == "merged":
                in_use = False
                for r in day_records:
                    if r["room"] not in ("01A", "01B"):
                        continue
                    op_s = to_minutes(r["start"])
                    op_e = to_minutes(r["end"])
                    if op_s <= iv_end and iv_start <= op_e:
                        in_use = True
                        break
                if in_use:
                    total += 1.0

            elif room_mode == "separate":
                for rm01 in ("01A", "01B"):
                    in_use = False
                    for r in day_records:
                        if r["room"] != rm01:
                            continue
                        op_s = to_minutes(r["start"])
                        op_e = to_minutes(r["end"])
                        if op_s <= iv_end and iv_start <= op_e:
                            in_use = True
                            break
                    if in_use:
                        total += 1.0

            elif room_mode == "01A_only":
                in_use = False
                for r in day_records:
                    if r["room"] != "01A":
                        continue
                    op_s = to_minutes(r["start"])
                    op_e = to_minutes(r["end"])
                    if op_s <= iv_end and iv_start <= op_e:
                        in_use = True
                        break
                if in_use:
                    total += 1.0

    n_intervals = len(interval_list)
    denom = num_rooms * n_intervals * num_days
    rate = total / denom * 100 if denom > 0 else 0
    diff_denom = denom - TARGET_DENOM
    diff_usage = total - TARGET_USAGE
    diff_rate = rate - TARGET_RATE
    return {
        "label": label,
        "usage": total,
        "denom": denom,
        "rate": rate,
        "n_intervals": n_intervals,
        "num_rooms": num_rooms,
        "diff_denom": diff_denom,
        "diff_usage": diff_usage,
        "diff_rate": diff_rate,
    }


# 区間パターン生成
def make_intervals(start_min, end_min, step):
    """[start_min, start_min+step-1], [start_min+step, ...] ..."""
    ivs = []
    m = start_min
    while m + step - 1 <= end_min:
        ivs.append((m, m + step - 1))
        m += step
    return ivs


# 9:00-16:59 = 540-1019 → 160区間
iv_160 = make_intervals(540, 1019, 3)
# 9:00-17:02 = 540-1022 → 161区間
iv_161 = make_intervals(540, 1022, 3)
# 8:57-16:59 = 537-1019 → 161区間
iv_161b = make_intervals(537, 1019, 3)
# 9:00-17:00 を161区間に（最後の区間が17:00-17:00の1分？）→ 別方式
# 8:59-16:59 → ずらし
iv_161c = make_intervals(539, 1022, 3)  # 8:59-17:02
# 9:00-17:00 = 540-1020 の場合、3分で割り切れない(481分)
# 別方式: 最初を9:00-9:02(3分), 最後を17:00-17:02(3分)含めて161区間

print(f"\n区間パターン確認:")
print(f"  iv_160: {len(iv_160)}区間, {iv_160[0]}~{iv_160[-1]} (9:00-16:59)")
print(f"  iv_161: {len(iv_161)}区間, {iv_161[0]}~{iv_161[-1]} (9:00-17:02)")
print(f"  iv_161b: {len(iv_161b)}区間, {iv_161b[0]}~{iv_161b[-1]} (8:57-16:59)")
print(f"  iv_161c: {len(iv_161c)}区間, {iv_161c[0]}~{iv_161c[-1]} (8:59-17:02)")

results = []

# === 仮説A: 1A/1B統合 + 9室 + 161区間(9:00-17:02) ===
results.append(calc_hypothesis("A: merged+9室+161区間(9:00-17:02)", "merged", 9, iv_161))

# === 仮説B: 1A/1B統合 + 9室 + 160区間(9:00-16:59) ===
results.append(calc_hypothesis("B: merged+9室+160区間(9:00-16:59)", "merged", 9, iv_160))

# === 仮説C: weighted(弊社) + 9室 + 161区間(9:00-17:02) ===
results.append(calc_hypothesis("C: weighted+9室+161区間(9:00-17:02)", "weighted", 9, iv_161))

# === 仮説D: weighted(弊社) + 9室 + 160区間(9:00-16:59) ===
results.append(calc_hypothesis("D: weighted+9室+160区間(弊社方式)", "weighted", 9, iv_160))

# === 仮説E: 1A/1B統合 + 9室 + 161区間(8:57-16:59) ===
results.append(calc_hypothesis("E: merged+9室+161区間(8:57-16:59)", "merged", 9, iv_161b))

# === 仮説F: separate(各1.0) + 10室 ===
results.append(calc_hypothesis("F: separate+10室+160区間", "separate", 10, iv_160))

# === 仮説G: 01Aのみ + 9室 + 161区間 ===
results.append(calc_hypothesis("G: 01Aonly+9室+161区間(9:00-17:02)", "01A_only", 9, iv_161))

# === 仮説H: 01Aのみ + 9室 + 160区間 ===
results.append(calc_hypothesis("H: 01Aonly+9室+160区間(9:00-16:59)", "01A_only", 9, iv_160))

# === 仮説I: merged + 9室 + 161区間(8:59-17:02) ===
results.append(calc_hypothesis("I: merged+9室+161区間(8:59-17:02)", "merged", 9, iv_161c))

# 表示
print(f"\n{'仮説':<38s}  {'使用合計':>10s}  {'分母':>8s}  {'稼働率':>8s}  {'分母差':>8s}  {'使用差':>8s}  {'率差':>8s}")
print("-" * 110)
for r in results:
    match_mark = ""
    if r["diff_denom"] == 0:
        match_mark = " <-- 分母一致!"
    if abs(r["diff_usage"]) <= 1 and r["diff_denom"] == 0:
        match_mark = " *** 完全一致! ***"
    print(f"{r['label']:<38s}  {r['usage']:10.1f}  {r['denom']:8d}  {r['rate']:7.2f}%  {r['diff_denom']:+8d}  {r['diff_usage']:+10.1f}  {r['diff_rate']:+7.2f}%{match_mark}")

print(f"\n他ソフト目標値: 使用={TARGET_USAGE}, 分母={TARGET_DENOM}, 稼働率={TARGET_RATE}%")

# ========== 3. 分母一致した仮説の詳細分析 ==========
print("\n" + "=" * 60)
print("3. 分母一致仮説の詳細分析")
print("=" * 60)
for r in results:
    if r["diff_denom"] == 0:
        print(f"\n--- {r['label']} ---")
        print(f"  分母: {r['denom']} = {r['num_rooms']}室 x {r['n_intervals']}区間 x {num_days}日 (一致)")
        print(f"  使用合計: {r['usage']:.1f} (目標{TARGET_USAGE}, 差={r['diff_usage']:+.1f})")
        print(f"  稼働率: {r['rate']:.2f}% (目標{TARGET_RATE}%, 差={r['diff_rate']:+.2f}pp)")
        if abs(r['diff_usage']) > 0:
            print(f"  使用合計の乖離: {abs(r['diff_usage']):.1f} ({abs(r['diff_usage'])/TARGET_USAGE*100:.2f}%)")

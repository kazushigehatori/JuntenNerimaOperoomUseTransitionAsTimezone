"""HOGY区間(0/+29) 全手術 分数ベース稼働率 01A+01B合算 分母10室 9:00-17:00 試行"""
import openpyxl
import datetime as dt

INPUT_FILE = "時間帯別稼働推移元データ.xlsx"
wb = openpyxl.load_workbook(INPUT_FILE)

def to_minutes(t):
    if isinstance(t, dt.time):
        return t.hour * 60 + t.minute
    elif isinstance(t, dt.timedelta):
        return int(t.total_seconds()) // 60
    elif isinstance(t, str):
        parts = t.split(":")
        return int(parts[0]) * 60 + int(parts[1])
    else:
        return t.hour * 60 + t.minute

# 定義シート
ws_def = wb["定義"]
room_weight = {}
for row in ws_def.iter_rows(min_row=2, max_row=20, min_col=1, max_col=2, values_only=True):
    if row[0] is not None and row[1] is not None:
        try:
            room_weight[str(row[0])] = float(row[1])
        except:
            break

exclude_weekdays = set()
r = 14
while True:
    r += 1
    v = ws_def.cell(row=r, column=1).value
    if v is None or v == "":
        break
    exclude_weekdays.add(str(v))

# データ読み込み
ws_data = wb["時間帯別稼働推移元データ"]
records = []
for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, values_only=True):
    mgmt_no, op_date, weekday, room, start_time, end_time, category = row
    if room is not None:
        records.append({
            "date": str(op_date) if op_date else "",
            "weekday": str(weekday) if weekday else "",
            "room": str(room),
            "start": start_time,
            "end": end_time,
            "category": str(category) if category else "",
        })

# フィルタ: 除外曜日除去、対象室のみ、全手術
filtered = [r for r in records if r["weekday"] not in exclude_weekdays
            and r["room"] in room_weight]

print(f"対象レコード数（全手術・除外曜日除去・対象室）: {len(filtered)}")
print(f"対象手術室: {list(room_weight.keys())}")
print(f"除外曜日: {exclude_weekdays}")

# HOGY区間: 0/+29 => [snap, snap+29] = 30分間
# スナップショット: 9:00〜16:30 (16区間) → 最終区間は16:30-16:59
INTERVAL_MINUTES = 30
snapshots = []
for h in range(9, 17):
    snapshots.append(dt.time(h, 0))
    snapshots.append(dt.time(h, 30))
print(f"スナップショット数: {len(snapshots)}")
print(f"スナップショット: {[f'{s.hour}:{s.minute:02d}' for s in snapshots]}")

# 分母=9室
DENOM_ROOMS = 9

# 日別にグループ化
days = {}
for r in filtered:
    d = r["date"]
    if d not in days:
        days[d] = []
    days[d].append(r)

num_days = len(days)
print(f"対象日数: {num_days}")

# 部屋一覧（01A+01Bは合算して1室扱い）
# 通常部屋: 02,03,05,06,07,08,09,10
# 合算部屋: 01（01A+01Bの重なり分をunionして算出）
normal_rooms = ["02", "03", "05", "06", "07", "08", "09", "10"]

# 各スナップショットの使用分数ベース稼働率
totals = [0.0] * len(snapshots)

for day_str, day_records in days.items():
    for si, snap in enumerate(snapshots):
        snap_min = to_minutes(snap)
        iv_start = snap_min           # 区間開始（含む）
        iv_end   = snap_min + 29      # 区間終了（含む）
        # 区間は iv_start 〜 iv_end の30分間

        day_usage = 0.0

        # --- 通常部屋（各1室）: 重なり分数を合算、上限30分 ---
        for rm in normal_rooms:
            rm_minutes = 0
            for r in day_records:
                if r["room"] != rm:
                    continue
                op_start = to_minutes(r["start"])
                op_end   = to_minutes(r["end"])
                # 重なり = [max(iv_start, op_start), min(iv_end, op_end)]
                overlap_start = max(iv_start, op_start)
                overlap_end   = min(iv_end, op_end)
                if overlap_start <= overlap_end:
                    rm_minutes += (overlap_end - overlap_start + 1)
            # 1部屋の上限=30分
            rm_minutes = min(rm_minutes, INTERVAL_MINUTES)
            day_usage += rm_minutes / INTERVAL_MINUTES

        # --- 01A, 01B 各ウェイト0.5（各部屋上限30分、ウェイト0.5）---
        for rm01 in ("01A", "01B"):
            rm01_minutes = 0
            for r in day_records:
                if r["room"] != rm01:
                    continue
                op_start = to_minutes(r["start"])
                op_end   = to_minutes(r["end"])
                overlap_start = max(iv_start, op_start)
                overlap_end   = min(iv_end, op_end)
                if overlap_start <= overlap_end:
                    rm01_minutes += (overlap_end - overlap_start + 1)
            rm01_minutes = min(rm01_minutes, INTERVAL_MINUTES)
            day_usage += (rm01_minutes / INTERVAL_MINUTES) * 0.5

        totals[si] += day_usage

averages = [t / num_days for t in totals]

print(f"\n=== HOGY区間(0/+29) / 全手術 / 分数ベース / 01A,01B各0.5 / 分母9室 ===")
print(f"{'時刻':>6s}  {'区間':>12s}  {'日平均使用室':>10s}  {'稼働率':>8s}")
print("-" * 46)
grand_total = 0.0
for si, snap in enumerate(snapshots):
    avg = averages[si]
    rate = avg / DENOM_ROOMS * 100
    grand_total += avg
    snap_min = to_minutes(snap)
    iv_label = f"{snap_min//60}:{snap_min%60:02d}-{(snap_min+29)//60}:{(snap_min+29)%60:02d}"
    print(f"{snap.hour:2d}:{snap.minute:02d}   {iv_label:>12s}  {avg:10.4f}  {rate:7.2f}%")

overall_avg = grand_total / len(snapshots)
overall_rate = overall_avg / DENOM_ROOMS * 100
print("-" * 46)
print(f"{'全体平均':>6s}                {overall_avg:10.4f}  {overall_rate:7.2f}%")

target = 67.9
diff = overall_rate - target
print(f"\n=== 67.9%との比較 ===")
print(f"算出稼働率: {overall_rate:.2f}%")
print(f"目標値:     {target:.1f}%")
print(f"差:         {diff:+.2f}pp")

print(f"\n※ 計算方法: 各部屋の区間内使用分数/30分 を稼働率とし、")
print(f"  01A,01Bは各ウェイト0.5（各上限30分×0.5=最大0.5）、通常部屋は各1.0")
print(f"  合計 / 分母9 で算出")
